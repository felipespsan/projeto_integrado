from flask import Flask, render_template, request, redirect, url_for, session, jsonify
from flask_login import LoginManager, UserMixin
import sqlite3

from openpyxl import load_workbook
import os

app = Flask(__name__)
app.secret_key = 'your_secret_key'

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = '/login'

class User(UserMixin):
    def __init__(self, id, username):
        self.id = id
        self.username = username

@login_manager.user_loader
def load_user(user_id):
    return User(user_id, get_user_by_username(user_id)["usuario"])

def check_credentials(username, password):
    conn = sqlite3.connect('bdusuarios.db')
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM usuarios WHERE usuario=? AND senha=?", (username, password))
    user = cursor.fetchone()
    conn.close()
    return user

def get_user_by_username(username):
    conn = sqlite3.connect("bdusuarios.db")
    cursor = conn.cursor()
    cursor.execute("SELECT nome, permissoes FROM usuarios WHERE usuario=?", (username,))
    user = cursor.fetchone()
    conn.close()
    if user:
        return {
            "nome": user[0],
            "permissoes": user[1]
        }
    return None

app.jinja_env.globals['get_user_by_username'] = get_user_by_username

def insert_user(username, password):
    conn = sqlite3.connect("bdusuarios.db")
    cursor = conn.cursor()
    cursor.execute("INSERT INTO usuarios (usuario, senha) VALUES (?, ?)", (username, password))
    conn.commit()
    success = cursor.rowcount > 0
    conn.close()
    return success

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/index', methods=['POST'])
def login():
    username = request.form['username']
    password = request.form['password']
    user = check_credentials(username, password)

    if user:
        session['username'] = username
        session.permanent = True
        return redirect(url_for('home'))
    else:
        return "<script>alert('Nome de usuário ou senha incorretos.'); window.location = '/login';</script>"


@app.route('/home')
def home():
    if 'username' in session:
        username = session['username']
        return render_template('index.html', username=username, message='Login efetuado com sucesso!')
    else:
        return redirect(url_for('index'))

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('index'))


@app.route('/login')
def login_page():
    return render_template('login.html')

@app.route('/cadastrar')
def cadastrar():
    return render_template('cadastrar.html')

@app.route('/ensaio')
def ensaio_page():
    if 'username' in session:
        return render_template('ensaio.html', username=session['username'])
    else:
        return redirect(url_for('index'))

@app.route('/calculos')
def calculos():
    if 'username' in session:
        return render_template('calculos.html', username=session['username'])
    else:
        return redirect(url_for('index'))

@app.route('/60811_1_1')
def nm60811_1_1():
    if 'username' in session:
        return render_template('60811_1_1.html', username=session['username'])
    else:
        return redirect(url_for('index'))

@app.route('/60811_1_2')
def nm60811_1_2():
    if 'username' in session:
        return render_template('60811_1_2.html', username=session['username'])
    else:
        return redirect(url_for('index'))
    
@app.route('/247_1')
def nm247_1():
    if 'username' in session:
        return render_template('247_1.html', username=session['username'])
    else:
        return redirect(url_for('index'))

#usuários
def get_users():
    conn = sqlite3.connect('bdusuarios.db')
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM usuarios")
    users = cursor.fetchall()
    conn.close()
    return users

@app.route('/users')
def users():
    if 'username' in session:
        users = get_users()
        return render_template('user_list.html', username=session['username'], users=users)
    else:
        return redirect(url_for('index'))


def get_user(username):
    conn = sqlite3.connect('bdusuarios.db')
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM usuarios WHERE usuario=?", (username,))
    user = cursor.fetchone()
    conn.close()
    return user


@app.route('/edit-user/<username>')
def user(username):
    if 'username' in session:
        user = get_user(username)
        return render_template('edit_user.html', user=user, username=session['username'])
    else:
        return redirect(url_for('index'))


@app.route('/update-user/<username>', methods=['POST'])
def update_user(username):
    try:
        username_new = request.form.get('username')
        password = request.form.get('senha')
        name = request.form.get('nome')
        gender = request.form.get('genero')
        permissions = request.form.get('permissoes')

        conn = sqlite3.connect('bdusuarios.db')
        cursor = conn.cursor()
        cursor.execute(
            "UPDATE usuarios SET usuario = ?, senha = ?, nome = ?, genero = ?, permissoes = ? WHERE usuario = ?",
            (username_new, password, name, gender, permissions, username)
        )

        conn.commit()
        conn.close()
    except Exception as e:
        return str(e)

    if 'username' in session:
        return redirect(url_for('users'))
    else:
        return redirect(url_for('index'))


@app.route('/realizar_cadastro', methods=['POST'])
def realizar_cadastro():
    username = request.form.get('username')
    password = request.form.get('password')
    nome = request.form.get('nome')
    genero = request.form.get('genero')

    success = insert_user(username, password, nome, genero)
    
    if success:
        if 'username' not in session:
            session['username'] = username
            session.permanent = True

        return redirect(url_for('home'))
    else:
        return "Erro ao cadastrar o usuário."


def insert_user(username, password, nome, genero):
    conn = sqlite3.connect("bdusuarios.db")
    cursor = conn.cursor()
    cursor.execute("INSERT INTO usuarios (usuario, senha, nome, genero, permissoes) VALUES (?, ?, ?, ?, 'usuário')",
                   (username, password, nome, genero))
    conn.commit()
    success = cursor.rowcount > 0
    conn.close()
    return success


@app.route('/delete-user/<username>', methods=['POST'])
def delete_user(username):
    success = delete_user_from_database(username)
    
    if success:
        return redirect(url_for('users'))
    else:
        return "Erro ao excluir o usuário."


def delete_user_from_database(username):
    conn = sqlite3.connect('bdusuarios.db')
    cursor = conn.cursor()
    cursor.execute("DELETE FROM usuarios WHERE usuario = ?", (username,))
    conn.commit()
    conn.close()

    return redirect(url_for('users'))

@app.errorhandler(400)
def handle_bad_request(e):
    return jsonify(error=str(e)), 400

def get_logged_in_user():
    if 'username' in session:
        return session['username']
    else:
        return None

@app.route('/salvar_informacoes', methods=['POST'])
def salvar_informacoes():
    coleta = request.form.get('coleta')
    usuario = session['username'] if 'username' in session else None
    composto = request.form.get('composto')
    distancia = request.form.get('distancia').replace('.', ',')
    minimo_alongamento = request.form.get('minimoAlongamento').replace('.', ',')
    minimo_ruptura = request.form.get('minimoRuptura')
    variacao_especificada = request.form.get('variacaoEspecificada')
    espessura1 = request.form.get('espessura1').replace('.', ',')
    espessura2 = request.form.get('espessura2').replace('.', ',')
    media_espessura = request.form.get('media_espessura')
    largura1 = request.form.get('largura1').replace('.', ',')
    largura2 = request.form.get('largura2').replace('.', ',')
    media_largura = request.form.get('media_largura')
    area1 = request.form.get('area1').replace('.', ',')
    area2 = request.form.get('area2').replace('.', ',')
    media_area = request.form.get('media_area')
    alongamento1 = request.form.get('alongamento1').replace('.', ',')
    alongamento2 = request.form.get('alongamento2').replace('.', ',')
    ruptura1 = request.form.get('ruptura1').replace('.', ',')
    ruptura2 = request.form.get('ruptura2').replace('.', ',')
    calculado1 = request.form.get('calculado1').replace('.', ',')
    calculado2 = request.form.get('calculado2').replace('.', ',')
    mediana_calculada = request.form.get('mediana_calculada')
    ruptura_area1 = request.form.get('ruptura_area1')
    ruptura_area2 = request.form.get('ruptura_area2')
    mediana_ruptura_area = request.form.get('mediana_ruptura_area')
    ruptura_MPa1 = request.form.get('ruptura_MPa1')
    ruptura_MPa2 = request.form.get('ruptura_MPa2')
    mediana_ruptura_MPa = request.form.get('mediana_ruptura_MPa')
    espessura1_2 = request.form.get('espessura1_2').replace('.', ',')
    espessura2_2 = request.form.get('espessura2_2').replace('.', ',')
    media_espessura_2 = request.form.get('media_espessura_2')
    largura1_2 = request.form.get('largura1_2').replace('.', ',')
    largura2_2 = request.form.get('largura2_2').replace('.', ',')
    media_largura_2 = request.form.get('media_largura_2')
    area1_2 = request.form.get('area1_2').replace('.', ',')
    area2_2 = request.form.get('area2_2').replace('.', ',')
    media_area_2 = request.form.get('media_area_2')
    alongamento1_2 = request.form.get('alongamento1_2').replace('.', ',')
    alongamento2_2 = request.form.get('alongamento2_2').replace('.', ',')
    ruptura1_2 = request.form.get('ruptura1_2').replace('.', ',')
    ruptura2_2 = request.form.get('ruptura2_2').replace('.', ',')
    calculado1_2 = request.form.get('calculado1_2').replace('.', ',')
    calculado2_2 = request.form.get('calculado2_2').replace('.', ',')
    mediana_calculada_2 = request.form.get('mediana_calculada_2')
    ruptura_area1_2 = request.form.get('ruptura_area1_2')
    ruptura_area2_2 = request.form.get('ruptura_area2_2')
    mediana_ruptura_area_2 = request.form.get('mediana_ruptura_area_2')
    ruptura_MPa1_2 = request.form.get('ruptura_MPa1_2')
    ruptura_MPa2_2 = request.form.get('ruptura_MPa2_2')
    mediana_ruptura_MPa_2 = request.form.get('mediana_ruptura_MPa_2')
    variacaoCalculada = request.form.get('variacaoCalculada')
    variacaoRupturaArea = request.form.get('variacaoRupturaArea')
    variacaoRupturaMPa = request.form.get('variacaoRupturaMPa')
    tipoCorpoDeProva = request.form.get('tipoCorpoDeProva')

    conn = sqlite3.connect('ensaio.db')
    cursor = conn.cursor()

    try:
        cursor.execute('SELECT coleta FROM informacoes WHERE coleta = ?', (coleta,))
        existing_coleta = cursor.fetchone()

        if existing_coleta:
            cursor.execute('''
                UPDATE informacoes SET
                    usuario = ?,
                    composto = ?,
                    distancia = ?,
                    minimoAlongamento = ?,
                    minimoRuptura = ?,
                    variacaoEspecificada = ?,
                    espessura1 = ?,
                    espessura2 = ?,
                    media_espessura = ?,
                    largura1 = ?,
                    largura2 = ?,
                    media_largura = ?,
                    area1 = ?,
                    area2 = ?,
                    media_area = ?,
                    alongamento1 = ?,
                    alongamento2 = ?,
                    ruptura1 = ?,
                    ruptura2 = ?,
                    calculado1 = ?,
                    calculado2 = ?,
                    mediana_calculada = ?,
                    ruptura_area1 = ?,
                    ruptura_area2 = ?,
                    mediana_ruptura_area = ?,
                    ruptura_MPa1 = ?,
                    ruptura_MPa2 = ?,
                    mediana_ruptura_MPa = ?,
                    espessura1_2 = ?,
                    espessura2_2 = ?,
                    media_espessura_2 = ?,
                    largura1_2 = ?,
                    largura2_2 = ?,
                    media_largura_2 = ?,
                    area1_2 = ?,
                    area2_2 = ?,
                    media_area_2 = ?,
                    alongamento1_2 = ?,
                    alongamento2_2 = ?,
                    ruptura1_2 = ?,
                    ruptura2_2 = ?,
                    calculado1_2 = ?,
                    calculado2_2 = ?,
                    mediana_calculada_2 = ?,
                    ruptura_area1_2 = ?,
                    ruptura_area2_2 = ?,
                    mediana_ruptura_area_2 = ?,
                    ruptura_MPa1_2 = ?,
                    ruptura_MPa2_2 = ?,
                    mediana_ruptura_MPa_2 = ?,
                    variacaoCalculada = ?,
                    variacaoRupturaArea = ?,
                    variacaoRupturaMPa = ?,
                    tipoCorpoDeProva = ?
                WHERE coleta = ?
            ''', (
                usuario, composto, distancia, minimo_alongamento, minimo_ruptura, variacao_especificada,
                espessura1, espessura2, media_espessura,
                largura1, largura2, media_largura,
                area1, area2, media_area,
                alongamento1, alongamento2,
                ruptura1, ruptura2,
                calculado1, calculado2, mediana_calculada,
                ruptura_area1, ruptura_area2, mediana_ruptura_area,
                ruptura_MPa1, ruptura_MPa2, mediana_ruptura_MPa,
                espessura1_2, espessura2_2, media_espessura_2,
                largura1_2, largura2_2, media_largura_2,
                area1_2, area2_2, media_area_2,
                alongamento1_2, alongamento2_2,
                ruptura1_2, ruptura2_2,
                calculado1_2, calculado2_2, mediana_calculada_2,
                ruptura_area1_2, ruptura_area2_2, mediana_ruptura_area_2,
                ruptura_MPa1_2, ruptura_MPa2_2, mediana_ruptura_MPa_2,
                variacaoCalculada, variacaoRupturaArea, variacaoRupturaMPa, tipoCorpoDeProva,
                coleta
            ))

        else:
            cursor.execute('''
                INSERT INTO informacoes (
                    coleta, usuario, composto, distancia, minimoAlongamento, minimoRuptura, variacaoEspecificada,
                    espessura1, espessura2, media_espessura,
                    largura1, largura2, media_largura,
                    area1, area2, media_area,
                    alongamento1, alongamento2,
                    ruptura1, ruptura2,
                    calculado1, calculado2, mediana_calculada,
                    ruptura_area1, ruptura_area2, mediana_ruptura_area,
                    ruptura_MPa1, ruptura_MPa2, mediana_ruptura_MPa,
                    espessura1_2, espessura2_2, media_espessura_2,
                    largura1_2, largura2_2, media_largura_2,
                    area1_2, area2_2, media_area_2,
                    alongamento1_2, alongamento2_2,
                    ruptura1_2, ruptura2_2,
                    calculado1_2, calculado2_2, mediana_calculada_2,
                    ruptura_area1_2, ruptura_area2_2, mediana_ruptura_area_2,
                    ruptura_MPa1_2, ruptura_MPa2_2, mediana_ruptura_MPa_2,
                    variacaoCalculada, variacaoRupturaArea, variacaoRupturaMPa, tipoCorpoDeProva
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                coleta, usuario, composto, distancia, minimo_alongamento, minimo_ruptura, variacao_especificada,
                espessura1, espessura2, media_espessura,
                largura1, largura2, media_largura,
                area1, area2, media_area,
                alongamento1, alongamento2,
                ruptura1, ruptura2,
                calculado1, calculado2, mediana_calculada,
                ruptura_area1, ruptura_area2, mediana_ruptura_area,
                ruptura_MPa1, ruptura_MPa2, mediana_ruptura_MPa,
                espessura1_2, espessura2_2, media_espessura_2,
                largura1_2, largura2_2, media_largura_2,
                area1_2, area2_2, media_area_2,
                alongamento1_2, alongamento2_2,
                ruptura1_2, ruptura2_2,
                calculado1_2, calculado2_2, mediana_calculada_2,
                ruptura_area1_2, ruptura_area2_2, mediana_ruptura_area_2,
                ruptura_MPa1_2, ruptura_MPa2_2, mediana_ruptura_MPa_2,
                variacaoCalculada, variacaoRupturaArea, variacaoRupturaMPa, tipoCorpoDeProva
            ))

        conn.commit()

    except Exception as e:
        return jsonify(error=str(e)), 400

    conn.close()

    return render_template('ensaio.html', username=session['username'], message='Informações salvas com sucesso!')

@app.route('/buscar_coleta')
def buscar_coleta():
    coleta = request.args.get('coleta')

    conn = sqlite3.connect('ensaio.db')
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM informacoes WHERE coleta = ?', (coleta,))
    data = cursor.fetchone()
    conn.close()

    if data:
        response = {
            'composto': data[2],
            'distancia': data[3],
            'minimoAlongamento': data[4],
            'minimoRuptura': data[5],
            'variacaoEspecificada': data[6],
            'espessura1': data[7],
            'espessura2': data[8],
            'media_espessura': data[9],
            'largura1': data[10],
            'largura2': data[11],
            'media_largura': data[12],
            'area1': data[13],
            'area2': data[14],
            'media_area': data[15],
            'alongamento1': data[16],
            'alongamento2': data[17],
            'ruptura1': data[18],
            'ruptura2': data[19],
            'calculado1': data[20],
            'calculado2': data[21],
            'mediana_calculada': data[22],
            'ruptura_area1': data[23],
            'ruptura_area2': data[24],
            'mediana_ruptura_area': data[25],
            'ruptura_MPa1': data[26],
            'ruptura_MPa2': data[27],
            'mediana_ruptura_MPa': data[28],
            'espessura1_2': data[29],
            'espessura2_2': data[30],
            'media_espessura_2': data[31],
            'largura1_2': data[32],
            'largura2_2': data[33],
            'media_largura_2': data[34],
            'area1_2': data[35],
            'area2_2': data[36],
            'media_area_2': data[37],
            'alongamento1_2': data[38],
            'alongamento2_2': data[39],
            'ruptura1_2': data[40],
            'ruptura2_2': data[41],
            'calculado1_2': data[42],
            'calculado2_2': data[43],
            'mediana_calculada_2': data[44],
            'ruptura_area1_2': data[45],
            'ruptura_area2_2': data[46],
            'mediana_ruptura_area_2': data[47],
            'ruptura_MPa1_2': data[48],
            'ruptura_MPa2_2': data[49],
            'mediana_ruptura_MPa_2': data[50],
            'variacaoCalculada': data[51],
            'variacaoRupturaArea': data[52],
            'variacaoRupturaMPa': data[53],
            'tipoCorpoDeProva': data[54]
        }
    else:
        response = {}

    return jsonify(response)

@app.route('/exportar_para_excel', methods=['POST'])
def exportar_para_excel():

    coleta = request.form.get('coleta')

    conn = sqlite3.connect('ensaio.db')
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM informacoes WHERE coleta = ?', (coleta,))
    data = cursor.fetchone()
    conn.close()

    if data is not None:
        coleta = data[0]
        usuario = data[1]
        composto = data[2]
        distancia = data[3]
        minimoAlongamento = data[4]
        minimoRuptura = data[5]
        variacaoEspecificada = data[6]
        espessura1 = data[7]
        espessura2 = data[8]
        media_espessura = data[9]
        largura1 = data[10]
        largura2 = data[11]
        media_largura = data[12]
        area1 = data[13]
        area2 = data[14]
        media_area = data[15]
        alongamento1 = data[16]
        alongamento2 = data[17]
        ruptura1 = data[18]
        ruptura2 = data[19]
        calculado1 = data[20]
        calculado2 = data[21]
        mediana_calculada = data[22]
        ruptura_area1 = data[23]
        ruptura_area2 = data[24]
        mediana_ruptura_area = data[25]
        ruptura_MPa1 = data[26]
        ruptura_MPa2 = data[27]
        mediana_ruptura_MPa = data[28]
        espessura1_2 = data[29]
        espessura2_2 = data[30]
        media_espessura_2 = data[31]
        largura1_2 = data[32]
        largura2_2 = data[33]
        media_largura_2 = data[34]
        area1_2 = data[35]
        area2_2 = data[36]
        media_area_2 = data[37]
        alongamento1_2 = data[38]
        alongamento2_2 = data[39]
        ruptura1_2 = data[40]
        ruptura2_2 = data[41]
        calculado1_2 = data[42]
        calculado2_2 = data[43]
        mediana_calculada_2 = data[44]
        ruptura_area1_2 = data[45]
        ruptura_area2_2 = data[46]
        mediana_ruptura_area_2 = data[47]
        ruptura_MPa1_2 = data[48]
        ruptura_MPa2_2 = data[49]
        mediana_ruptura_MPa_2 = data[50]
        variacaoCalculada = data[51]
        variacaoRupturaArea = data[52]
        variacaoRupturaMPa = data[53]
        tipoCorpoDeProva = data[54]

        workbook = load_workbook("Coletas.xlsx")

        sheet = workbook["dados"]

        sheet.cell(row=2, column=1).value = coleta
        sheet.cell(row=2, column=2).value = usuario
        sheet.cell(row=2, column=3).value = composto
        sheet.cell(row=2, column=4).value = distancia
        sheet.cell(row=2, column=5).value = minimoAlongamento
        sheet.cell(row=2, column=6).value = minimoRuptura
        sheet.cell(row=2, column=7).value = variacaoEspecificada
        sheet.cell(row=2, column=8).value = espessura1
        sheet.cell(row=2, column=9).value = espessura2
        sheet.cell(row=2, column=10).value = media_espessura
        sheet.cell(row=2, column=11).value = largura1
        sheet.cell(row=2, column=12).value = largura2
        sheet.cell(row=2, column=13).value = media_largura
        sheet.cell(row=2, column=14).value = area1
        sheet.cell(row=2, column=15).value = area2
        sheet.cell(row=2, column=16).value = media_area
        sheet.cell(row=2, column=17).value = alongamento1
        sheet.cell(row=2, column=18).value = alongamento2
        sheet.cell(row=2, column=19).value = ruptura1
        sheet.cell(row=2, column=20).value = ruptura2
        sheet.cell(row=2, column=21).value = calculado1
        sheet.cell(row=2, column=22).value = calculado2
        sheet.cell(row=2, column=23).value = mediana_calculada
        sheet.cell(row=2, column=24).value = ruptura_area1
        sheet.cell(row=2, column=25).value = ruptura_area2
        sheet.cell(row=2, column=26).value = mediana_ruptura_area
        sheet.cell(row=2, column=27).value = ruptura_MPa1
        sheet.cell(row=2, column=28).value = ruptura_MPa2
        sheet.cell(row=2, column=29).value = mediana_ruptura_MPa
        sheet.cell(row=2, column=30).value = espessura1_2
        sheet.cell(row=2, column=31).value = espessura2_2
        sheet.cell(row=2, column=32).value = media_espessura_2
        sheet.cell(row=2, column=33).value = largura1_2
        sheet.cell(row=2, column=34).value = largura2_2
        sheet.cell(row=2, column=35).value = media_largura_2
        sheet.cell(row=2, column=36).value = area1_2
        sheet.cell(row=2, column=37).value = area2_2
        sheet.cell(row=2, column=38).value = media_area_2
        sheet.cell(row=2, column=39).value = alongamento1_2
        sheet.cell(row=2, column=40).value = alongamento2_2
        sheet.cell(row=2, column=41).value = ruptura1_2
        sheet.cell(row=2, column=42).value = ruptura2_2
        sheet.cell(row=2, column=43).value = calculado1_2
        sheet.cell(row=2, column=44).value = calculado2_2
        sheet.cell(row=2, column=45).value = mediana_calculada_2
        sheet.cell(row=2, column=46).value = ruptura_area1_2
        sheet.cell(row=2, column=47).value = ruptura_area2_2
        sheet.cell(row=2, column=48).value = mediana_ruptura_area_2
        sheet.cell(row=2, column=49).value = ruptura_MPa1_2
        sheet.cell(row=2, column=50).value = ruptura_MPa2_2
        sheet.cell(row=2, column=51).value = mediana_ruptura_MPa_2
        sheet.cell(row=2, column=52).value = variacaoCalculada
        sheet.cell(row=2, column=53).value = variacaoRupturaArea
        sheet.cell(row=2, column=54).value = variacaoRupturaMPa
        sheet.cell(row=2, column=55).value = tipoCorpoDeProva

        workbook.save("Coletas.xlsx")

        os.startfile("Coletas.xlsx")

        return "Relatório exportado com sucesso!", 200


import webbrowser
webbrowser.open_new('http://localhost:5000/')

if __name__ == '__main__':
    #app.debug = True
    app.run(host='0.0.0.0')